import streamlit as st
import pandas as pd
import numpy as np
from pandasai import PandasAI
from pandasai.llm.openai import OpenAI
from PIL import Image
import matplotlib.pyplot as plt
import openai
import openpyxl

#EXL logo
image = Image.open('exl.png')
df = pd.read_csv('data.csv')


with st.sidebar:
	st.image(image, width = 150)
	st.write('Ask any question on your data')

#define tabs
tab1, tab2, tab3 = st.tabs(['Analysis','Report', 'Validation'])

ls = ['chart','plot','graph','trend', 'histogram']
#to check if prompt have chart, graph words
def contains_substring(string, substrings):
	for substring in substrings:
		if substring in string:
			return True
	return False


# tab1 for raw data analysis
with tab1:
	
	llm = OpenAI(api_token=st.secrets["chat_gpt_key"])
	
	pandas_ai = PandasAI(llm, conversational=False)#, enforce_privacy = True)

	
	st.header("Conversational BI")
	st.subheader("Sample data structure ")
	st.dataframe(df.head())
	
	with st.form("conversation_bi"):
	
		query = st.text_input(label ="Enter a question" , placeholder = 'Enter your query')
		# Every form must have a submit button.
		submitted1 = st.form_submit_button("Submit")
		if submitted1:
			#based on type of response
			if contains_substring(query.lower(),ls): 
				fig, x = plt.subplots()
				response1 = pandas_ai(df, prompt=query)
				st.pyplot(fig)
			else:
				response1 = pandas_ai(df, prompt=query)
				if isinstance(response, str):
					st.text(response1)
				elif isinstance(response1, pd.DataFrame):
					st.dataframe(response1)
				else:
					st.text(response1.to_string(index=False))
				


with tab2:

	df2 = pd.read_csv('report.csv')
	col1, col2 = st.columns(2)
	
	with col1:
		#pie chart for sales by category
		total_sales = df2.groupby('Category')['Sales'].sum()
		fig1, ax1= plt.subplots()
		ax1.pie(total_sales.values, labels=total_sales.index, autopct='%1.1f%%')
		ax1.set_title('Sales by Category')
		st.pyplot(fig1)
		
		#pie chart for sales by region
		total_sales2 = df2.groupby('Region')['Sales'].sum()
		fig3, ax3= plt.subplots()
		ax3.pie(total_sales2.values, labels=total_sales2.index, autopct='%1.1f%%')
		ax3.set_title('Sales by Region')
		st.pyplot(fig3)
		
	with col2:	
		#average sales by category
		average_sales = df2.groupby('Category')['Sales'].mean().reset_index()
		fig2, ax2= plt.subplots()
		ax2.bar(average_sales['Category'], average_sales['Sales'])
		ax2.set_title('Average Sales by Region')
		plt.xticks(rotation=45)
		st.pyplot(fig2)
		
		#average sales by region
		average_sales2 = df2.groupby('Region')['Sales'].mean().reset_index()
		fig4, ax4= plt.subplots()
		ax4.bar(average_sales2['Region'], average_sales2['Sales'])
		ax4.set_title('Average Sales by Region')
		plt.xticks(rotation=45)
		st.pyplot(fig4)

	
	with st.form("select category"):
		select_option = st.multiselect('Please select categories', df2['Category'].unique())
		generate_mails = st.form_submit_button("Generate Communication")
		if generate_mails:
			st.write('Go to next tab to validate communication')

with tab3:
			
	st.header("Personalized communication ")
	path = "data-mail.xlsx"
	wb= openpyxl.load_workbook(path)
	ws = wb.active

		
	with st.form("communication"):
		df3 = df2[df2['Category'].isin(select_option)]
		name = st.selectbox('Please select agent to check outgoing communication',df3["Name"])
		category = df2[df2.Name == name]['Category'].to_string(index=False)
		target = df2[df2.Name == name]['Target'].to_string(index=False)
		curr_sales = df2[df2.Name == name]['Sales'].to_string(index=False)
		
		# row and column index of required obs
		name_index = ws.cell(row = df.loc[df.Name == name].index[0] + 2
		mail_index = 10
		# Every form must have a submit button.
		submitted2 = st.form_submit_button("Validate")
		if submitted2:
			st.text(f"""Name: {name}\nCategory : {category}\nTarget : ${target}\nCurrnt Sales : ${curr_sales}""")
			st.write()
			mail_response = ws.cell(row = name_index , column = mail_index).value
			st.markdown(mail_response)
	#if "mail_response" in globals():
	#	st.write('Please select name first')
	#else:
	if st.button('Edit', key = 'ABC'):
		if not name:
			st.write('Please select name first')
		else:
			with st.form("edit communication"):
				name_index = ws.cell(row = df.loc[df.Name == name].index[0] + 2
				mail_index = 10
				user_input = st.text_area("Edit the mail",height = 600, value= ws.cell(row = name_index , column = mail_index).value)
				submitted3 = st.form_submit_button("Save")
				if submitted3:
					ws.cell(row = name_index , column = mail_index).value = user_input
					wb.save('path') 
					
