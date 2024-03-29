import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from openpyxl.drawing.image import Image as ii


#EXL logo
image = Image.open('exl.png')
#read data file in dataframe
#df = pd.read_csv('data.csv')
df = pd.read_excel('Mort_V1.xlsx', header=2)


with st.sidebar:
	st.image(image, width = 150)
	st.write('Ask any question on your data')


def pivot1(df,ind, col):
	#Pivot the data and calculate the total actual and expected deaths for each product and duration 
	pivot_table = pd.pivot_table(df, values=['Actual Deaths','Expected Deaths'], index=ind, columns = col, aggfunc='sum', margins=True, margins_name="Total")
	
	#Calculate the percentage of total actual deaths compared to total expected deaths 
	for i in df[col].unique():
		pivot_table['Percentage',i] = (pivot_table['Actual Deaths',i]/ pivot_table['Expected Deaths',i]* 100).round().map('{:.0f}%'.format)
	pivot_table['Percentage','Total'] = (pivot_table['Actual Deaths','Total']/ pivot_table['Expected Deaths','Total']* 100).round().map('{:.0f}%'.format)
	
	# Display the pivot table with both headers
	
	#Add a title to the pivot table 
	#title = "Percentage of Total Actual Deaths vs Total Expected Deaths for Different Products and Durations" 
	#pivot_table_with_title = pd.concat([pd.DataFrame([title], columns=['']), pivot_table], axis=0)
	df1 = pivot_table['Percentage']
	df1.reset_index(level=0, inplace=True)
	df1.columns.values[0] = ind + '/'+ col
	return df1


def pivot2(df,col):
	df1 = df.groupby(col)[['Actual Deaths','Expected Deaths']].sum()
	df1['Mortality'] = (df1['Actual Deaths']/df1['Expected Deaths'] * 100).round().map('{:.0f}%'.format)
	df1.reset_index(level=0, inplace=True)
	return df1
	
# function to add value labels
def addlabels(x,y):
    for i in range(len(x)):
        plt.text(i, y[i]+0.5, str(y[i])+'%', ha = 'center')

st.subheader("Conversational BI")
st.write("Sample data structure ")
st.dataframe(df.head())
#st.markdown(df.style.hide(axis="index").to_html(), unsafe_allow_html=True)

#with st.form("conversation_bi"):

inp_query = st.text_input(label ="Enter a question" , placeholder = 'Enter your query')
query = inp_query.lower()
chart2 = 'n'

if st.button("Submit"):

	if query == 'show mortality experience analysis by product and duration':
		df_out = pivot1(df,'Product', 'Duration')
		title = 'Mortality experience by Product and Duration'

		df_t = df_out.set_index('Product/Duration').reset_index()
		df_t['Total'] = pd.to_numeric(df_t['Total'].str.strip('%').replace('nan',0))
		x_labels = df_t['Product/Duration'].tolist()
		
		# Set the positions of the bars on the x-axis
		x = range(len(x_labels))
		# Plot the bar graph for each value of Term and Endowment
		fig, ax = plt.subplots()
		bar_width = 0.5
		plt.bar(x, df_t['Total'], width=bar_width)
	
		# Label the axes and add a title
		ax.set_xlabel('Product')
		ax.set_xticks([pos for pos in x])
		ax.set_xticklabels(x_labels)
		ax.set_ylabel('Mortality Experience')
		ax.set_title('Mortality experience by Product')
		addlabels(x_labels, df_t['Total'])
		fig.savefig('Graph1.png')
		
	elif query == 'show mortality experience analysis by product and smoker status':
		df_out = pivot1(df,'Product', 'Smoker Status')
		title = 'Mortality experience by Product and Smoker Status'

		df_t = df_out.set_index('Product/Smoker Status').T.reset_index()
		#df_t = df_t[df_t['Smoker Status'] != 'Total']
		df_t['Total'] = pd.to_numeric(df_t['Total'].str.strip('%').replace('nan',0))
		x_labels = df_t['Smoker Status'].tolist()
		# Set the positions of the bars on the x-axis
		x = range(len(x_labels))
		# Plot the bar graph for each value of Term and Endowment
		fig, ax = plt.subplots()
		bar_width = 0.35
		plt.bar(x, df_t['Total'], width=bar_width)
		# Label the axes and add a title
		ax.set_xlabel('Smoker Status')
		ax.set_xticks([pos for pos in x])
		ax.set_xticklabels(x_labels)
		ax.set_ylabel('Mortality Experience')
		ax.set_title('Mortality Experience by Smoker Status')
		addlabels(x_labels, df_t['Total'])
		fig.savefig('Graph1.png')
		
		chart2 = 'y'
		df_t2 = df_out.set_index('Product/Smoker Status').reset_index()
		df_t2['Total'] = pd.to_numeric(df_t2['Total'].str.strip('%').replace('nan',0))
		x_labels = df_t2['Product/Smoker Status'].tolist()
		# Set the positions of the bars on the x-axis
		x = range(len(x_labels))
		# Plot the bar graph for each value of Term and Endowment
		fig2, ax2 = plt.subplots()
		bar_width = 0.35
		plt.bar(x, df_t2['Total'], width=bar_width)
	
		# Label the axes and add a title
		ax2.set_xlabel('Product')
		ax2.set_xticks([pos for pos in x])
		ax2.set_xticklabels(x_labels)
		ax2.set_ylabel('Mortality Experience')
		ax2.set_title('Mortality Experience by Product')
		addlabels(x_labels, df_t2['Total'])
		fig2.savefig('Graph2.png')
		
	elif query == 'show mortality experience analysis by sum assured class and product':
		df_out = pivot1(df,'Sum Assured Class', 'Product')
		title = 'Mortality experience by Sum Assured Class and Product'

		df_t = df_out.set_index('Sum Assured Class/Product').reset_index()
		#df_t = df_t[df_t['Sum Assured Class/Product'] != 'Total']
		df_t['Total'] = pd.to_numeric(df_t['Total'].str.strip('%').replace('nan',0))
		x_labels = df_t['Sum Assured Class/Product'].tolist()
		# Set the positions of the bars on the x-axis
		x = range(len(x_labels))
		# Plot the bar graph for each value of Term and Endowment
		fig, ax = plt.subplots()
		bar_width = 0.5
		plt.bar(x, df_t['Total'], width=bar_width)
		# Label the axes and add a title
		ax.set_xlabel('Sum Assured Class')
		ax.set_xticks([pos for pos in x])
		ax.set_xticklabels(x_labels)
		ax.set_ylabel('Mortality Experience')
		ax.set_title('Mortality Experience by Sum Assured Class')
		addlabels(x_labels, df_t['Total'])
		fig.savefig('Graph1.png')
		
		chart2 = 'y'
		df_t2 = df_out.set_index('Sum Assured Class/Product').T.reset_index()
		#df_t = df_t[df_t['Sum Assured Class/Product'] != 'Total']
		df_t2['Total'] = pd.to_numeric(df_t2['Total'].str.strip('%').replace('nan',0))
		x_labels = df_t2['Product'].tolist()
		# Set the positions of the bars on the x-axis
		x = range(len(x_labels))
		# Plot the bar graph for each value of Term and Endowment
		fig2, ax2 = plt.subplots()
		bar_width = 0.5
		plt.bar(x, df_t2['Total'], width=bar_width)
		# Label the axes and add a title
		ax2.set_xlabel('Product')
		ax2.set_xticks([pos for pos in x])
		ax2.set_xticklabels(x_labels)
		ax2.set_ylabel('Mortality Experience')
		ax2.set_title('Mortality Experience by Product')
		addlabels(x_labels, df_t2['Total'])
		fig2.savefig('Graph2.png')
		
	elif query == 'show mortality experience analysis by issue year':
		df_out = pivot2(df,'Issue Year')
		df_out['Issue Year'] = df_out['Issue Year'].astype(str).str.replace(',', '')
		title = 'Mortality experience by Issue Year'

		fig, ax = plt.subplots()
		x = pd.to_numeric(df_out['Issue Year'].str.strip('%').replace('nan',0))
		y = pd.to_numeric(df_out['Mortality'].str.strip('%').replace('nan',0))
		plt.plot(x, y, marker='o', linestyle='-')
		ax.set_xlabel('Year')
		ax.set_ylabel('Mortality Experience')
		ax.set_title(title)
		fig.savefig('Graph1.png')
		
	elif query == 'show mortality experience analysis by uw class':
		df_out = pivot2(df,'UW Class')
		title = 'Mortality experience by UW Class'
		
		fig, ax = plt.subplots()
		x = df_out['UW Class']
		y = pd.to_numeric(df_out['Mortality'].str.strip('%').replace('nan',0))
		bar_width = 0.35
		plt.bar(x, y, width=bar_width)
		ax.set_xlabel('UW Class')
		ax.set_ylabel('Mortality Experience')
		ax.set_title(title)
		addlabels(x,y)
		fig.savefig('Graph1.png')

	st.write(title)
	st.dataframe(df_out)
	#st.markdown(df_out.style.hide(axis="index").to_html(), unsafe_allow_html=True)

	col1, col2 = st.columns(2)
	with col1:
		st.pyplot(fig)
	
		#Excel
		workbook = Workbook()
		sheet = workbook.active
		sheet.title = 'Report'
	
		c1 = sheet.cell(row = 1, column = 1)
		c1.value = title
		for row in dataframe_to_rows(df_out, index = False):
			sheet.append(row)
	
		#formatting graph
		sheet2 = workbook.create_sheet(title='Graph')
		graph = ii('Graph1.png')
		aspect_ratio = graph.width / graph.height
		graph.width = 600
		graph.height = graph.width/aspect_ratio
		#adding graph to sheet
		sheet2.add_image(graph, 'B2')
	with col2:
		if chart2 == 'y':
			st.pyplot(fig2)
			#formatting graph
			graph2 = ii('Graph2.png')
			aspect_ratio2 = graph2.width / graph2.height
			graph2.width = 600
			graph2.height = graph2.width/aspect_ratio2
			#adding graph to sheet
			sheet2.add_image(graph2, 'L2')

		workbook.save('output.xlsx')

	
	with open("output.xlsx", "rb") as file:
		st.download_button(
			label="Download Excel",
			data=file,
			file_name='data.xlsx'
		)

		
